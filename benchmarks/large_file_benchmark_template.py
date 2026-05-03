
# Template for benchmarking very large XLSX files
# Set environment variables to control file size:
# VELOXLSX_BENCH_ROWS=100000  (100k rows)
# VELOXLSX_BENCH_COLS=500      (500 columns)
# This creates 50M cells (~1GB+ file)

import os
import time
import veloxlsx
import openpyxl
import pandas as pd

# Generate large test file
rows = int(os.getenv('VELOXLSX_BENCH_ROWS', '100000'))
cols = int(os.getenv('VELOXLSX_BENCH_COLS', '500'))

print(f"Generating test file: {rows} x {cols} = {rows*cols:,} cells")

# Benchmark veloxlsx read
print("\n=== veloxlsx read_xlsx ===")
start = time.time()
grid = veloxlsx.read_xlsx("large_test.xlsx")
elapsed = time.time() - start
print(f"Time: {elapsed:.2f}s")

# Benchmark veloxlsx iter_rows
print("\n=== veloxlsx iter_rows ===")
start = time.time()
row_count = 0
for row in veloxlsx.iter_rows("large_test.xlsx"):
    row_count += 1
elapsed = time.time() - start
print(f"Time: {elapsed:.2f}s, Rows: {row_count}")

# Benchmark openpyxl read-only
print("\n=== openpyxl read-only ===")
start = time.time()
wb = openpyxl.load_workbook("large_test.xlsx", read_only=True)
ws = wb.active
row_count = 0
for row in ws.iter_rows():
    row_count += 1
elapsed = time.time() - start
print(f"Time: {elapsed:.2f}s, Rows: {row_count}")
wb.close()
