"""
Run in a fresh interpreter (subprocess) so `resource.ru_maxrss` reflects that workload.

Usage:
  python memory_worker.py read <path>
  python memory_worker.py load_read <path>
  python memory_worker.py openpyxl_read <path>
  python memory_worker.py calamine_read <path>
  python memory_worker.py pandas_calamine_read <path>
  python memory_worker.py pandas_openpyxl_read <path>
  python memory_worker.py stream_write <path> <rows> <cols>
  python memory_worker.py write_xlsx <path> <rows> <cols>
  python memory_worker.py xlsxwriter_write <path> <rows> <cols>
"""

from __future__ import annotations

import json
import resource
import sys
import time


def peak_rss_mib() -> float:
    rss = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
    if sys.platform == "darwin":
        return rss / (1024.0 * 1024.0)
    # Linux: kilobytes
    return rss / 1024.0


def main() -> None:
    mode = sys.argv[1]
    if mode == "read":
        path = sys.argv[2]
        import fast_xlsx

        t0 = time.perf_counter()
        fast_xlsx.read_xlsx(path)
        ms = (time.perf_counter() - t0) * 1000.0
    elif mode == "load_read":
        path = sys.argv[2]
        import fast_xlsx

        t0 = time.perf_counter()
        wb = fast_xlsx.load(path)
        wb.read_sheet(0)
        ms = (time.perf_counter() - t0) * 1000.0
    elif mode == "openpyxl_read":
        path = sys.argv[2]
        from openpyxl import load_workbook

        t0 = time.perf_counter()
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        assert ws is not None
        for _ in ws.iter_rows(values_only=True):
            pass
        wb.close()
        ms = (time.perf_counter() - t0) * 1000.0
    elif mode == "calamine_read":
        path = sys.argv[2]
        from python_calamine import load_workbook

        t0 = time.perf_counter()
        wb = load_workbook(str(path))
        wb.get_sheet_by_index(0).to_python()
        ms = (time.perf_counter() - t0) * 1000.0
    elif mode == "pandas_calamine_read":
        path = sys.argv[2]
        import pandas as pd

        t0 = time.perf_counter()
        df = pd.read_excel(path, header=None, engine="calamine")
        df.to_numpy(copy=False)
        ms = (time.perf_counter() - t0) * 1000.0
    elif mode == "pandas_openpyxl_read":
        path = sys.argv[2]
        import pandas as pd

        t0 = time.perf_counter()
        df = pd.read_excel(path, header=None, engine="openpyxl")
        df.to_numpy(copy=False)
        ms = (time.perf_counter() - t0) * 1000.0
    elif mode == "stream_write":
        path = sys.argv[2]
        rows = int(sys.argv[3])
        cols = int(sys.argv[4])
        import fast_xlsx

        t0 = time.perf_counter()
        with fast_xlsx.StreamWriter(path, sheet_name="Sheet1") as w:
            for r in range(rows):
                w.write_row([r * 1_000_000 + c for c in range(cols)])
        ms = (time.perf_counter() - t0) * 1000.0
    elif mode == "write_xlsx":
        path = sys.argv[2]
        rows = int(sys.argv[3])
        cols = int(sys.argv[4])
        import fast_xlsx

        grid = [[r * 1_000_000 + c for c in range(cols)] for r in range(rows)]
        t0 = time.perf_counter()
        fast_xlsx.write_xlsx(path, grid, sheet="Sheet1")
        ms = (time.perf_counter() - t0) * 1000.0
    elif mode == "xlsxwriter_write":
        path = sys.argv[2]
        rows = int(sys.argv[3])
        cols = int(sys.argv[4])
        import xlsxwriter

        t0 = time.perf_counter()
        wb = xlsxwriter.Workbook(path, {"constant_memory": True})
        ws = wb.add_worksheet("Sheet1")
        for r in range(rows):
            ws.write_row(r, 0, [r * 1_000_000 + c for c in range(cols)])
        wb.close()
        ms = (time.perf_counter() - t0) * 1000.0
    else:
        raise SystemExit(f"unknown mode: {mode}")

    out = {"ms": ms, "peak_rss_mib": peak_rss_mib()}
    print(json.dumps(out))


if __name__ == "__main__":
    main()
