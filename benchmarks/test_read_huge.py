"""
Read benchmarks on a large numeric grid.

Run from repo root (after `maturin develop`):

  pytest benchmarks/

Tune size with env vars (defaults: 4000 x 120 = 480k cells):

  VELOXLSX_BENCH_ROWS=8000 VELOXLSX_BENCH_COLS=200 pytest benchmarks/

CI uses smaller values via workflow env for a quick smoke run.
"""

from __future__ import annotations

from pathlib import Path

import pytest


@pytest.mark.benchmark(group="read_huge")
def test_bench_veloxlsx_huge(benchmark, huge_xlsx_path: Path) -> None:
    import veloxlsx

    benchmark(lambda: veloxlsx.read_xlsx(huge_xlsx_path))


@pytest.mark.benchmark(group="read_huge")
def test_bench_openpyxl_read_only_huge(benchmark, huge_xlsx_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    def run() -> None:
        wb = load_workbook(huge_xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        assert ws is not None
        for _ in ws.iter_rows(values_only=True):
            pass
        wb.close()

    benchmark(run)


@pytest.mark.benchmark(group="read_huge")
def test_bench_python_calamine_huge(benchmark, huge_xlsx_path: Path) -> None:
    pytest.importorskip("python_calamine")
    from python_calamine import load_workbook

    def run() -> None:
        wb = load_workbook(str(huge_xlsx_path))
        wb.get_sheet_by_index(0).to_python()

    benchmark(run)


@pytest.mark.benchmark(group="read_huge")
def test_bench_pandas_calamine_huge(benchmark, huge_xlsx_path: Path) -> None:
    pytest.importorskip("pandas")
    import pandas as pd

    def run() -> None:
        df = pd.read_excel(huge_xlsx_path, header=None, engine="calamine")
        df.to_numpy(copy=False)

    benchmark(run)


@pytest.mark.benchmark(group="read_huge")
def test_bench_pandas_openpyxl_huge(benchmark, huge_xlsx_path: Path) -> None:
    pytest.importorskip("pandas")
    pytest.importorskip("openpyxl")
    import pandas as pd

    def run() -> None:
        df = pd.read_excel(huge_xlsx_path, header=None, engine="openpyxl")
        df.to_numpy(copy=False)

    benchmark(run)


def test_huge_fixture_dimensions(huge_xlsx_path: Path, huge_cell_count: int) -> None:
    """Non-benchmark sanity check so `pytest benchmarks/` always does something useful."""
    assert huge_xlsx_path.is_file()
    assert huge_cell_count >= 50 * 10
