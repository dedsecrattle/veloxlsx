"""
Generate visual comparison graphs for benchmark results.
Run this after running memory_timing.py to visualize the results.
"""

import matplotlib.pyplot as plt
import numpy as np
from pathlib import Path


def create_comparison_charts():
    """Create comparison charts for library performance."""
    
    # Benchmark data from README (4000 × 120 numeric grid, ~480k cells)
    # These are sample results - replace with actual benchmark data
    
    # Read performance data
    read_data = {
        'Library': [
            'veloxlsx\nread_xlsx',
            'veloxlsx\niter_rows',
            'veloxlsx\nload+read_sheet',
            'openpyxl\niter_rows',
            'python-\ncalamine',
            'pandas\ncalamine',
            'pandas\nopenpyxl'
        ],
        'Time (ms)': [236.3, 258.8, 241.0, 603.4, 196.0, 228.2, 812.6],
        'Memory (MiB)': [114.7, 36.0, 114.6, 38.9, 68.9, 141.4, 101.5]
    }
    
    # Write performance data
    write_data = {
        'Library': [
            'veloxlsx\nStreamWriter',
            'veloxlsx\nwrite_xlsx',
            'XlsxWriter\nconstant_memory'
        ],
        'Time (ms)': [298.3, 273.0, 829.9],
        'Memory (MiB)': [15.6, 70.7, 24.3]
    }
    
    # Create figure with subplots
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('XLSX Library Performance Comparison\n(4000 × 120 grid, ~480k cells)', 
                 fontsize=16, fontweight='bold')
    
    # Read Time Comparison
    ax1 = axes[0, 0]
    colors_read_time = ['#2e86ab', '#2e86ab', '#2e86ab', '#f26419', '#8cb369', '#f8961e', '#f94144']
    bars1 = ax1.bar(read_data['Library'], read_data['Time (ms)'], color=colors_read_time)
    ax1.set_ylabel('Time (ms)', fontsize=12, fontweight='bold')
    ax1.set_title('Read Performance - Time', fontsize=13, fontweight='bold')
    ax1.tick_params(axis='x', rotation=45)
    ax1.grid(axis='y', alpha=0.3)
    
    # Add value labels on bars
    for bar in bars1:
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2., height,
                f'{height:.1f}',
                ha='center', va='bottom', fontsize=9)
    
    # Read Memory Comparison
    ax2 = axes[0, 1]
    colors_read_mem = ['#2e86ab', '#2e86ab', '#2e86ab', '#f26419', '#8cb369', '#f8961e', '#f94144']
    bars2 = ax2.bar(read_data['Library'], read_data['Memory (MiB)'], color=colors_read_mem)
    ax2.set_ylabel('Peak RSS (MiB)', fontsize=12, fontweight='bold')
    ax2.set_title('Read Performance - Memory', fontsize=13, fontweight='bold')
    ax2.tick_params(axis='x', rotation=45)
    ax2.grid(axis='y', alpha=0.3)
    
    # Add value labels on bars
    for bar in bars2:
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2., height,
                f'{height:.1f}',
                ha='center', va='bottom', fontsize=9)
    
    # Write Time Comparison
    ax3 = axes[1, 0]
    colors_write_time = ['#2e86ab', '#2e86ab', '#f26419']
    bars3 = ax3.bar(write_data['Library'], write_data['Time (ms)'], color=colors_write_time)
    ax3.set_ylabel('Time (ms)', fontsize=12, fontweight='bold')
    ax3.set_title('Write Performance - Time', fontsize=13, fontweight='bold')
    ax3.tick_params(axis='x', rotation=45)
    ax3.grid(axis='y', alpha=0.3)
    
    # Add value labels on bars
    for bar in bars3:
        height = bar.get_height()
        ax3.text(bar.get_x() + bar.get_width()/2., height,
                f'{height:.1f}',
                ha='center', va='bottom', fontsize=9)
    
    # Write Memory Comparison
    ax4 = axes[1, 1]
    colors_write_mem = ['#2e86ab', '#2e86ab', '#f26419']
    bars4 = ax4.bar(write_data['Library'], write_data['Memory (MiB)'], color=colors_write_mem)
    ax4.set_ylabel('Peak RSS (MiB)', fontsize=12, fontweight='bold')
    ax4.set_title('Write Performance - Memory', fontsize=13, fontweight='bold')
    ax4.tick_params(axis='x', rotation=45)
    ax4.grid(axis='y', alpha=0.3)
    
    # Add value labels on bars
    for bar in bars4:
        height = bar.get_height()
        ax4.text(bar.get_x() + bar.get_width()/2., height,
                f'{height:.1f}',
                ha='center', va='bottom', fontsize=9)
    
    plt.tight_layout()
    
    # Save the figure
    output_path = Path(__file__).parent / 'benchmark_comparison.png'
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    print(f"Benchmark comparison chart saved to: {output_path}")
    
    # Create a scatter plot showing time vs memory tradeoff
    fig2, ax = plt.subplots(figsize=(12, 8))
    
    # Read libraries scatter
    read_x = read_data['Memory (MiB)']
    read_y = read_data['Time (ms)']
    read_labels = read_data['Library']
    read_colors = ['#2e86ab' if 'veloxlsx' in lib else '#f26419' if 'openpyxl' in lib 
                   else '#8cb369' if 'calamine' in lib else '#f8961e' for lib in read_labels]
    
    ax.scatter(read_x, read_y, c=read_colors, s=200, alpha=0.7, edgecolors='black', linewidth=2)
    
    # Write libraries scatter
    write_x = write_data['Memory (MiB)']
    write_y = write_data['Time (ms)']
    write_labels = write_data['Library']
    write_colors = ['#2e86ab' if 'veloxlsx' in lib else '#f26419' for lib in write_labels]
    
    ax.scatter(write_x, write_y, c=write_colors, s=200, alpha=0.7, edgecolors='black', 
               linewidth=2, marker='s')
    
    # Add labels for each point
    for i, label in enumerate(read_labels):
        ax.annotate(label.replace('\n', ' '), (read_x[i], read_y[i]),
                   fontsize=9, ha='center', va='bottom')
    
    for i, label in enumerate(write_labels):
        ax.annotate(label.replace('\n', ' ') + ' (write)', (write_x[i], write_y[i]),
                   fontsize=9, ha='center', va='bottom')
    
    ax.set_xlabel('Peak Memory (MiB)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Time (ms)', fontsize=12, fontweight='bold')
    ax.set_title('Time vs Memory Tradeoff\n(Circle: Read, Square: Write)', 
                 fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3)
    
    # Add legend
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='#2e86ab', label='veloxlsx'),
        Patch(facecolor='#8cb369', label='python-calamine'),
        Patch(facecolor='#f26419', label='openpyxl/XlsxWriter'),
        Patch(facecolor='#f8961e', label='pandas')
    ]
    ax.legend(handles=legend_elements, loc='upper right')
    
    plt.tight_layout()
    
    # Save the scatter plot
    output_path2 = Path(__file__).parent / 'benchmark_scatter.png'
    plt.savefig(output_path2, dpi=300, bbox_inches='tight')
    print(f"Scatter plot saved to: {output_path2}")
    
    return output_path, output_path2


def create_large_file_benchmark_template():
    """Create a template script for benchmarking very large files."""
    
    template = """
# Template for benchmarking very large XLSX files
# Set environment variables to control file size:
# VELOXLSX_BENCH_ROWS=100000  (100k rows)
# VELOXLSX_BENCH_COLS=500      (500 columns)
# This creates 50M cells (~1GB+ file)

import os
import time
import veloxlsx
import openpyxl
import pandas as pd

# Generate large test file
rows = int(os.getenv('VELOXLSX_BENCH_ROWS', '100000'))
cols = int(os.getenv('VELOXLSX_BENCH_COLS', '500'))

print(f"Generating test file: {rows} x {cols} = {rows*cols:,} cells")

# Benchmark veloxlsx read
print("\\n=== veloxlsx read_xlsx ===")
start = time.time()
grid = veloxlsx.read_xlsx("large_test.xlsx")
elapsed = time.time() - start
print(f"Time: {elapsed:.2f}s")

# Benchmark veloxlsx iter_rows
print("\\n=== veloxlsx iter_rows ===")
start = time.time()
row_count = 0
for row in veloxlsx.iter_rows("large_test.xlsx"):
    row_count += 1
elapsed = time.time() - start
print(f"Time: {elapsed:.2f}s, Rows: {row_count}")

# Benchmark openpyxl read-only
print("\\n=== openpyxl read-only ===")
start = time.time()
wb = openpyxl.load_workbook("large_test.xlsx", read_only=True)
ws = wb.active
row_count = 0
for row in ws.iter_rows():
    row_count += 1
elapsed = time.time() - start
print(f"Time: {elapsed:.2f}s, Rows: {row_count}")
wb.close()
"""
    
    template_path = Path(__file__).parent / 'large_file_benchmark_template.py'
    with open(template_path, 'w') as f:
        f.write(template)
    
    print(f"Large file benchmark template saved to: {template_path}")
    return template_path


if __name__ == '__main__':
    print("Generating benchmark comparison charts...")
    chart_path, scatter_path = create_comparison_charts()
    
    print("\\nGenerating large file benchmark template...")
    template_path = create_large_file_benchmark_template()
    
    print("\\nDone! To benchmark very large files:")
    print("1. Generate a large file using: python benchmarks/huge_fixture.py")
    print("2. Set environment variables: VELOXLSX_BENCH_ROWS=100000 VELOXLSX_BENCH_COLS=500")
    print("3. Run: python benchmarks/memory_timing.py")
    print("4. Update the data in this script with your results")
    print("5. Re-run this script to generate new charts")
