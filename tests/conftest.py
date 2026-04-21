from __future__ import annotations

from pathlib import Path

import pytest


@pytest.fixture()
def sample_xlsx_path(tmp_path: Path) -> Path:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws["A1"] = "hello"
    ws["B1"] = 42.5
    ws["C1"] = True
    ws["A2"] = "shared"
    ws["B2"] = "shared"
    wb.save(path)
    return path


@pytest.fixture()
def inline_str_xlsx_path(tmp_path: Path) -> Path:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    path = tmp_path / "inline.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "inline text here"
    wb.save(path)
    return path


@pytest.fixture()
def large_xlsx_path(tmp_path: Path) -> Path:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    path = tmp_path / "large.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    rows = 200
    cols = 50
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * 1000 + c)
    wb.save(path)
    return path
