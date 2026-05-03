from __future__ import annotations

from pathlib import Path

import pytest


@pytest.mark.benchmark(group="read")
def test_benchmark_veloxlsx_read(benchmark, large_xlsx_path: Path) -> None:
    import veloxlsx

    benchmark(lambda: veloxlsx.read_xlsx(large_xlsx_path))


@pytest.mark.benchmark(group="read")
def test_benchmark_openpyxl_read(benchmark, large_xlsx_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    def run() -> None:
        wb = load_workbook(large_xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        assert ws is not None
        for _ in ws.iter_rows(values_only=True):
            pass
        wb.close()

    benchmark(run)
