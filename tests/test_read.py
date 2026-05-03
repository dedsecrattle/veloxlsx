from __future__ import annotations

from pathlib import Path

import pytest

import veloxlsx


def test_read_xlsx_first_sheet(sample_xlsx_path: Path) -> None:
    grid = veloxlsx.read_xlsx(sample_xlsx_path)
    assert grid[0][0] == "hello"
    assert grid[0][1] == 42.5
    assert grid[0][2] is True
    assert grid[1][0] == "shared"
    assert grid[1][1] == "shared"


def test_read_xlsx_by_sheet_name(sample_xlsx_path: Path) -> None:
    grid = veloxlsx.read_xlsx(sample_xlsx_path, "Data")
    assert grid[0][0] == "hello"


def test_read_xlsx_by_sheet_index(sample_xlsx_path: Path) -> None:
    assert veloxlsx.read_xlsx(sample_xlsx_path, 0)[0][0] == "hello"
    assert veloxlsx.read_xlsx(sample_xlsx_path, -1)[0][0] == "hello"


def test_load_workbook_and_sheet(sample_xlsx_path: Path) -> None:
    wb = veloxlsx.load(sample_xlsx_path)
    assert wb.sheet_names == ["Data"]
    assert wb.read_sheet("Data")[0][0] == "hello"
    sheet = wb["Data"]
    assert sheet.name == "Data"
    assert sheet.to_list()[1][0] == "shared"


def test_inline_string_roundtrip(inline_str_xlsx_path: Path) -> None:
    grid = veloxlsx.read_xlsx(inline_str_xlsx_path)
    assert grid[0][0] == "inline text here"


def test_openpyxl_parity(sample_xlsx_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")

    from openpyxl import load_workbook

    wb = load_workbook(sample_xlsx_path, data_only=True)
    ws = wb["Data"]
    rust = veloxlsx.read_xlsx(sample_xlsx_path, "Data")
    for r, row in enumerate(rust, start=1):
        for c, value in enumerate(row, start=1):
            expected = ws.cell(row=r, column=c).value
            assert value == expected
